import datetime
import enum
from math import ceil
from typing import override

from loguru import logger
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from src.core.alias import ProgramDay, ProgramPHD
from src.core.datetime_utils import date2str, time2str, addDelta2Time
from src.domain.file_output_name_use_case import FileOutputNameUseCase
from src.presentation.view import PHDProgramView

HEIGHT_HEADER = 4


class _DayPage:
    def __init__(
            self,
            title_page: str,
            workbook: Workbook,
            day: datetime.date,
            program: ProgramDay,
            create_new_list: bool = True
    ):
        self._wb = workbook
        self._day = day
        self._program_day = program
        if create_new_list:
            self._create_new_list()
        self._page.title = title_page

    @property
    def _page(self):
        return self._wb.active

    def present(self):
        raise NotImplementedError

    def _create_new_list(self):
        ws = self._wb.create_sheet()
        self._wb.active = self._wb.worksheets.index(ws)


class _SimpleDayPageColumn(enum.Enum):
    time = "A"
    name = "B"
    tag = "C"
    talk_type = "D"
    place = "E"
    description = "F"
    speakers = "G"
    short = "H"
    broadcast = "I"

    @staticmethod
    def first() -> '_SimpleDayPageColumn':
        return [*_SimpleDayPageColumn][0]

    @staticmethod
    def last() -> '_SimpleDayPageColumn':
        return [*_SimpleDayPageColumn][-1]

    @staticmethod
    def count() -> int:
        return len(_SimpleDayPageColumn)

    def __str__(self):
        return self.value


class _SimpleDayPage(_DayPage):

    @override
    def present(self):
        self.__tune_header_cells()
        self.__fill_page_day(self._day, self._program_day)

    def __fill_page_day(self, day: datetime.date, program_day: ProgramDay):
        self.__fill_header(day)
        self.__fill_content(program_day)

    def __tune_header_cells(self):
        for column in self._page.iter_rows(min_col=1, max_col=_SimpleDayPageColumn.count(), min_row=1,
                                           max_row=HEIGHT_HEADER - 1):
            for cell in column:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

    def __fill_content(self, program_day: ProgramDay):
        def fill_column(letter: str, index: int, value: str, is_center: bool = True):
            cell = self._page[f'{letter}{index}']
            cell.value = value
            if is_center:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        for i in range(len(program_day)):
            e = program_day[i]
            row = i + HEIGHT_HEADER
            fill_column(_SimpleDayPageColumn.time.value, row, e.time)
            fill_column(_SimpleDayPageColumn.tag.value, row, e.tag.name)
            fill_column(_SimpleDayPageColumn.name.value, row, e.title)
            fill_column(_SimpleDayPageColumn.place.value, row, e.map_object.title)
            fill_column(_SimpleDayPageColumn.speakers.value, row,
                        ", ".join([f"{i.name} ({i.company})" for i in e.speakers]))
            fill_column(_SimpleDayPageColumn.description.value, row,
                        e.description if e.description is not None else "-", False)
            fill_column(_SimpleDayPageColumn.talk_type.value, row, e.talk_type.name if e.talk_type is not None else "-")
            fill_column(_SimpleDayPageColumn.broadcast.value, row, "Есть" if e.is_visible_in_broadcast else "Нет")
            fill_column(_SimpleDayPageColumn.short.value, row, f"{e.time}\n{e.title}\n{e.map_object.title}")

    def __fill_header(self, day: datetime.date):
        self._page['A1'] = f"СОЗДАН {datetime.datetime.today().strftime('%d.%m.%Y')}"
        self._page['B1'] = day.strftime('PHD - %d.%m.%Y')
        self._page.merge_cells(f'{_SimpleDayPageColumn.first()}1:{_SimpleDayPageColumn.last()}1')

        def fill_header_column(column_latter: str, title: str, width: int | None):
            self._page[f'{column_latter}2'] = title
            self._page.merge_cells(f'{column_latter}2:{column_latter}3')
            if width is not None:
                self._page.column_dimensions[column_latter].width = width

        fill_header_column(_SimpleDayPageColumn.time.value, "Время", 13)
        fill_header_column(_SimpleDayPageColumn.name.value, "Наименование", 80)
        fill_header_column(_SimpleDayPageColumn.tag.value, "Трек", 30)
        fill_header_column(_SimpleDayPageColumn.talk_type.value, "Тип", 25)
        fill_header_column(_SimpleDayPageColumn.place.value, "Зал", 25)
        fill_header_column(_SimpleDayPageColumn.speakers.value, "Спикеры", 40)
        fill_header_column(_SimpleDayPageColumn.description.value, "Описание", 100)
        fill_header_column(_SimpleDayPageColumn.short.value, "SHORT", 40)
        fill_header_column(_SimpleDayPageColumn.broadcast.value, "Трансляция", 15)


class _TimeDayPage(_DayPage):

    @override
    def present(self):
        self.__fill_time_column()
        self.__fill_content()

    MIN_TIME = datetime.time(hour=9)
    MAX_TIME = datetime.time(hour=23, minute=59)
    MINUTES_PER_ROW = 5

    def __init__(self, day: datetime.date, title_page: str, workbook: Workbook, program: ProgramDay):
        super().__init__(title_page, workbook, day, program)
        self._max_total_minute = self.MAX_TIME.minute + self.MAX_TIME.hour * 60
        self._min_total_minute = self.MIN_TIME.minute + self.MIN_TIME.hour * 60

    def __calc_count_rows(self) -> int:
        total = self._max_total_minute - self._min_total_minute
        return ceil(total / self.MINUTES_PER_ROW)

    def __fill_time_column(self):
        time = self.MIN_TIME
        count_rows = self.__calc_count_rows()
        for cells in self._page.iter_rows(min_row=1, max_row=count_rows, max_col=0):
            cell = cells[0]
            cell.value = time2str(time)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            time = addDelta2Time(time, datetime.timedelta(minutes=self.MINUTES_PER_ROW))

    def __find_index_row_by_time(self, time: datetime.time) -> int:
        total_minute = time.minute + time.hour * 60 - self._min_total_minute
        if (total_minute % self.MINUTES_PER_ROW) != 0:
            raise IndexError(f"Time ({time}) should be even {self.MINUTES_PER_ROW}")
        return total_minute // self.MINUTES_PER_ROW + 1

    def __find_near_free_column_by_row(self, index_row: int) -> int:
        for cells in self._page.iter_rows(min_row=index_row, max_row=index_row):
            for cell in cells:
                if cell.value is None and not isinstance(cell, MergedCell):
                    return cell.column
        return self._page.max_column + 1

    def __fill_content(self):
        for report in self._program_day:
            try:
                start_index_row = self.__find_index_row_by_time(report.start_time)
                end_index_row = self.__find_index_row_by_time(report.end_time)
                index_column = self.__find_near_free_column_by_row(start_index_row)
                cell = self._page.cell(row=start_index_row, column=index_column)
                cell.value = f"{report.time}\n{report.title}\n{report.map_object.title}"
                self._page.merge_cells(
                    start_row=start_index_row, end_row=end_index_row,
                    start_column=index_column, end_column=index_column
                )
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                column_letter = get_column_letter(index_column)
                self._page.column_dimensions[column_letter].width = 15
                for i in range(start_index_row, end_index_row+1):
                    self._page.row_dimensions[i].height = 30
            except IndexError as e:
                logger.error(e)


class ExcelPHDProgramView(PHDProgramView):
    def __init__(self, file_output_use_case: FileOutputNameUseCase):
        super().__init__()
        self.__file_output_name_use_case = file_output_use_case
        self._wb = Workbook()

    @property
    def __page(self):
        return self._wb.active

    @override
    def present(self):
        logger.info("Presenting Excel View")
        for i, (day, program_day) in enumerate(self._phd_program.items()):
            simple_page = _SimpleDayPage(
                title_page=date2str(day),
                workbook=self._wb,
                program=program_day,
                day=day,
                create_new_list=i != len(self._phd_program) - 1
            )
            simple_page.present()
            time_page = _TimeDayPage(
                title_page=f"{date2str(day)} (по времени)",
                workbook=self._wb,
                program=program_day,
                day=day,
            )
            time_page.present()
        self.__save()

    def __save(self):
        name = self.__file_output_name_use_case.get_available_output_name(format_file="xlsx")
        logger.info(f"Saving Excel View with {name=}")
        self._wb.save(name)

    def __create_new_list(self):
        ws = self._wb.create_sheet()
        self._wb.active = self._wb.worksheets.index(ws)
